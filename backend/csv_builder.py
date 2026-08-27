"""
KISA 취약점 점검 결과 -> 스타일이 적용된 다중 시트 XLSX 보고서 생성
기존 generate_csv()를 대체/보완하는 generate_xlsx() 제공

입력 데이터 형식 (기존 generate_csv와 동일한 hosts_data 사용, 필드 추가는 선택):
hosts_data = [
    {
        "hostname": "rocky1",
        "ip": "192.168.1.10",
        "os": "Rocky Linux 9.2",          # 선택, 없으면 "-"
        "owner": "보안관리자",             # 선택, 없으면 "-"
        "security_score_100": 87.5,       # 선택, 없으면 결과 목록으로 양호율 계산해서 대체
        "results": [
            {
                "code": "U-01",
                "category": "계정 관리",
                "title": "root 계정 원격 접속 제한",
                "importance": "상",          # 상/중/하
                "status": "취약",            # 양호/취약/검토/N/A
                "fixed_by_user": 0,           # 선택: 실제 조치 스크립트로 고쳤는지
                "manual_verdict": "",         # 선택: 사람이 확정한 판정("양호"/"취약")
                "evidence_description": "...",   # 현재 설정 값 (상태와 무관하게 항상 표시)
                "recommendation_text": "...",    # 조치 내용 및 미조치시 사유
                "target_file": "/etc/ssh/sshd_config",  # 선택
            },
            ...
        ],
    },
    ...
]

보안점수는 여기서 대충 재계산하지 않고, 서버(recompute_host_score)가 이미
weight_score/manual_verdict를 반영해서 계산해둔 host["security_score_100"]을
그대로 쓴다 - 안 그러면 대시보드/서버 목록에 보이는 점수와 여기 나오는 점수가
서로 달라진다. 그 필드가 없는(문서 스키마만 따르는) 입력에 대해서만 단순
양호율로 폴백한다.
"""

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.worksheet.worksheet import Worksheet

# 서버(컨트롤 노드)의 시스템 타임존은 UTC라서 datetime.now()가 UTC 벽시계를 반환한다.
# 보고서 표지에 찍히는 기간/생성일은 KST로 명시해서 기록한다.
KST = ZoneInfo("Asia/Seoul")

# ---------------------------------------------------------------------------
# 공통 스타일 정의 (첨부 예시 파일에서 추출한 값 그대로 사용)
# ---------------------------------------------------------------------------
NAVY = "1E293B"
LIGHT_GRAY = "F8FAFC"
BORDER_GRAY = "E2E8F0"

STATUS_STYLE = {
    "양호": {"fill": "DCFCE7", "font": "166534"},
    "취약": {"fill": "FEE2E2", "font": "991B1B"},
    "검토": {"fill": "E0F2FE", "font": "075985"},
    "N/A": {"fill": "F1F5F9", "font": "475569"},
}
IMPORTANCE_STYLE = {
    "상": {"fill": "FFE4E6", "font": "9F1239"},
    "중": {"fill": "FEF3C7", "font": "92400E"},
    "하": {"fill": "E0F2FE", "font": "075985"},
}

thin = Side(style="thin", color=BORDER_GRAY)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FONT = Font(bold=True, size=9.5, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, size=20, color="0F172A")
SUBTITLE_FONT = Font(size=12, color="64748B")
LABEL_FONT = Font(bold=True, size=9, color="000000")
VALUE_FONT = Font(size=11, color="000000")
NAV_FONT = Font(size=9, color="0284C7", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _status_key(raw_status: str) -> str:
    """DB 상태를 양호/취약/검토/N/A 네 가지로 정규화한다."""
    if raw_status is None or str(raw_status).strip() == "":
        return "검토"
    value = str(raw_status).strip().upper()
    if value in ("양호", "OK", "GOOD", "PASS"):
        return "양호"
    if value in ("취약", "FAIL", "VULNERABLE"):
        return "취약"
    if value in ("N/A", "NA", "NOT APPLICABLE", "해당없음"):
        return "N/A"
    return "검토"  # 검토, 수동확인 등


def _effective_status(result: dict) -> str:
    """manual_verdict(관리자가 '검토' 항목을 양호/취약으로 확정한 값)가 있으면
    그걸 우선한다 - db.py의 recompute_host_score(점수 계산)가 이미 이 규칙을
    쓰는데, 보고서 통계/표는 원래 status만 보고 있어서 확정한 뒤에도 계속
    '검토'로 잡히는 불일치가 있었다. status 원본 컬럼 자체는 DB에서 안
    건드린다(자동 진단 결과 기록 보존) - 보고서에 보여줄 값만 여기서 맞춘다."""
    return _status_key(result.get("manual_verdict") or result.get("status"))


def _action_status(result: dict, status_key: str) -> str:
    """진단 판정과 실제 조치 상태를 혼동하지 않도록 별도로 표시한다.

    reviewed는 apply_remediation_result()가 "조치" 스크립트를 한 번이라도
    돌려 재검사했을 때(성공/실패 무관) 1로 세팅된다 - 이게 없으면 "한 번도
    안 건드린 취약"과 "조치를 시도했지만 안 고쳐진 취약"이 똑같이 "미조치"로
    보여서 재조치가 필요한 항목을 놓치기 쉽다."""
    if result.get("fixed_by_user"):
        return "조치 완료"
    if result.get("manual_verdict"):
        return "수동 판정"
    if status_key == "취약":
        return "재조치 필요" if result.get("reviewed") else "미조치"
    if status_key == "검토":
        return "검토 필요"
    return "조치 불필요"  # 양호 / N/A


def _wrap_row_height(texts, col_width, min_h=28, max_h=360):
    """wrap_text=True만으로는 openpyxl이 실제 줄바꿈 수만큼 행 높이를 늘려주지
    않는다(엑셀에서 직접 열어야 자동으로 늘어난다) - LibreOffice 등으로 그냥
    렌더링/변환하면 첫 줄만 보이고 잘린 것처럼 나온다. 셀에 들어갈 텍스트
    길이와 열 너비로 필요한 줄 수를 대략 추정해 행 높이를 미리 넉넉히 준다.
    한글은 라틴 문자보다 훨씬 넓어서(openpyxl의 열 너비 단위는 라틴 문자
    기준) 줄당 글자 수를 낮게 잡아야 실제 줄바꿈 수와 비슷해진다."""
    chars_per_line = max(int(col_width * 0.85), 8)
    max_lines = 1
    for t in texts:
        s = str(t or "")
        if not s:
            continue
        wrapped_lines = sum(-(-max(len(part), 1) // chars_per_line) for part in s.split("\n"))
        max_lines = max(max_lines, wrapped_lines)
    return min(max_h, max(min_h, max_lines * 15 + 12))


def _score(status_key: str) -> int:
    return 10 if status_key == "양호" else 0


def _host_score(host: dict) -> float:
    """호스트 보안점수. 서버가 계산해둔 security_score_100(가중치 +
    manual_verdict 반영)이 있으면 그대로 쓰고, 없을 때만(문서 스키마만 따르는
    호출부 등) 결과 목록의 단순 양호율로 대체한다."""
    stored = host.get("security_score_100")
    if stored is not None:
        return float(stored)
    results = host.get("results", [])
    if not results:
        return 0.0
    good = sum(_effective_status(r) == "양호" for r in results)
    return round((good / len(results)) * 100, 2)


def _style_header_row(ws: Worksheet, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER_ALL


def _nav_link(ws: Worksheet, cell_ref: str, text: str, target_sheet: str):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = NAV_FONT
    cell.hyperlink = f"#'{target_sheet}'!A1"


def _autofit(ws: Worksheet, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# 데이터 집계
# ---------------------------------------------------------------------------
def _aggregate(hosts_data):
    """전체 통계, 중요도별 통계, 영역별 통계, 코드별 통계를 미리 계산"""
    total = {"양호": 0, "취약": 0, "검토": 0, "N/A": 0}
    by_importance = {}  # 상/중/하 -> 상태별 건수
    by_category = {}  # 점검영역 -> 상태별 건수
    by_code = {}  # code -> {category, title, importance, 전체, 취약, 취약서버}

    for h in hosts_data:
        hostname = h.get("hostname", "-")
        for r in h.get("results", []):
            sk = _effective_status(r)
            total[sk] += 1

            imp = r.get("importance", "중")
            by_importance.setdefault(imp, {"양호": 0, "취약": 0, "검토": 0, "N/A": 0})
            by_importance[imp][sk] += 1

            cat = r.get("category", "기타")
            by_category.setdefault(cat, {"양호": 0, "취약": 0, "검토": 0, "N/A": 0})
            by_category[cat][sk] += 1

            code = r.get("code", "-")
            entry = by_code.setdefault(
                code,
                {
                    "category": cat,
                    "title": r.get("title", ""),
                    "importance": imp,
                    "전체": 0,
                    "취약": 0,
                    "취약서버": [],
                },
            )
            entry["전체"] += 1
            if sk == "취약":
                entry["취약"] += 1
                entry["취약서버"].append(hostname)

    return total, by_importance, by_category, by_code


# ---------------------------------------------------------------------------
# 시트별 생성 함수
# ---------------------------------------------------------------------------
def _build_cover(wb: Workbook, meta: dict):
    ws = wb.create_sheet("표지")
    ws.sheet_view.showGridLines = False
    _nav_link(ws, "A1", ">> 대시보드", "대시보드")

    # 상단 강조 바 - 그냥 텍스트만 나열하지 않고 색 블록으로 시작점을 표시
    for c in range(2, 9):
        ws.cell(9, c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[9].height = 6

    ws["B11"] = meta.get("title", "서버 취약점 진단 상세 결과 보고서")
    ws["B11"].font = TITLE_FONT
    ws.merge_cells("B11:H11")

    ws["B13"] = meta.get("subtitle", "UNIX·DBMS 통합 보안 진단")
    ws["B13"].font = SUBTITLE_FONT
    ws.merge_cells("B13:H13")

    # 정보 표 - 줄글 나열 대신 라벨/값 칸을 표로 구분해서 훑어보기 쉽게
    info_rows = [
        ("수행 기관", meta.get("org") or "HIGHFIVE SECURITY"),
        ("고객사", meta.get("customer", "-")),
        ("진단 기간", meta.get("period", "-")),
        ("대상 서버", f"{meta.get('host_count', 0)}대"),
        ("점검 항목", f"UNIX {meta.get('unix_count', 0)}개, DBMS {meta.get('db_count', 0)}개 (총 {meta.get('item_count', 0)}개)"),
        ("진단자", meta.get("auditor", "-")),
    ]
    # "진단자"(수행 인력 팀 명단)와는 별개로, 채워 넣은 경우에만 추가되는
    # 보고서 담당자 1명 - DOCX 표지와 동일한 이유로 팀 명단을 덮어쓰지 않는다.
    if meta.get("inspector"):
        info_rows.append(("진단 담당자", meta["inspector"]))
    start_row = 17
    for i, (label, value) in enumerate(info_rows):
        r = start_row + i
        lbl = ws.cell(r, 2, label); lbl.font = Font(bold=True, size=10, color="334155")
        lbl.fill = PatternFill("solid", fgColor=LIGHT_GRAY); lbl.border = BORDER_ALL
        lbl.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        val = ws.cell(r, 3, value); val.font = Font(size=10.5, color="0F172A")
        val.border = BORDER_ALL; val.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 20

    _autofit(ws, {"A": 3, "B": 16, "C": 14, "D": 11, "E": 11, "F": 11, "G": 20, "H": 25})
    return ws


def _build_dashboard(wb: Workbook, hosts_data, total, by_importance, by_category, by_code):
    ws = wb.create_sheet("대시보드")
    _nav_link(ws, "A1", "<< 표지", "표지")

    # 핵심 카드: 열자마자 우선 조치 대상을 인지하도록 배치
    host_scores = [(_host_score(h), h.get("hostname", "-"), h.get("ip", "-")) for h in hosts_data]
    critical = sum(1 for h in hosts_data for r in h.get("results", []) if _effective_status(r) == "취약" and r.get("importance") == "상")
    cards = [("대상 서버", f"{len(hosts_data)}대"), ("취약 항목", f"{total['취약']}건"), ("상 위험 취약", f"{critical}건"), ("검토 필요", f"{total['검토']}건")]
    for col, (label, value) in zip((2, 3, 4, 5), cards):
        a = ws.cell(3, col, label); b = ws.cell(4, col, value)
        for cell in (a, b):
            cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = CENTER; cell.border = BORDER_ALL
        a.font = Font(size=9, color="CBD5E1", bold=True); b.font = Font(size=16, color="FFFFFF", bold=True)
    ws.merge_cells("B6:E6"); ws["B6"] = "우선순위: 상 위험 취약점 → 보안 점수 하위 서버 → 반복 취약점 순으로 조치하세요."
    ws["B6"].font = Font(size=9, color="475569", italic=True)

    # 전체 결과 + 도넛 차트 (N/A는 "해당 없음"이라 진단 결과 비율에서 제외)
    for c, label in zip(("B10", "C10"), ("구분", "건수")):
        ws[c] = label; ws[c].fill = HEADER_FILL; ws[c].font = HEADER_FONT; ws[c].alignment = CENTER; ws[c].border = BORDER_ALL
    for i, label in enumerate(("양호", "취약", "검토"), start=11):
        ws.cell(i, 2, label); ws.cell(i, 3, total[label])
        style = STATUS_STYLE[label]
        ws.cell(i, 2).fill = PatternFill("solid", fgColor=style["fill"]); ws.cell(i, 2).font = Font(color=style["font"], bold=True)
        for c in (2, 3): ws.cell(i, c).border = BORDER_ALL; ws.cell(i, c).alignment = CENTER
    chart = DoughnutChart(); chart.title = "진단 결과 비율"
    chart.add_data(Reference(ws, min_col=3, min_row=10, max_row=13), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=2, min_row=11, max_row=13)); chart.height = 7; chart.width = 10; ws.add_chart(chart, "E10")

    # 점수 하위 서버
    ws["B17"] = "보안 점수 하위 서버"; ws["B17"].font = Font(bold=True, size=11)
    for c, label in enumerate(("순위", "서버", "IP", "점수"), start=2):
        cell = ws.cell(18, c, label); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = BORDER_ALL
    for i, (score, name, ip) in enumerate(sorted(host_scores)[:5], start=19):
        for c, v in enumerate((i - 18, name, ip, f"{score:.2f}점"), start=2):
            cell = ws.cell(i, c, v); cell.border = BORDER_ALL; cell.alignment = CENTER if c in (2, 5) else Alignment(vertical="center")
            if c == 5 and score < 60: cell.font = Font(color="991B1B", bold=True)

    # 반복 취약점 Top 5
    start_row = 26
    ws.cell(start_row, 2, "반복 취약점 Top 5").font = Font(bold=True, size=11)
    for c, label in enumerate(("코드", "항목명", "취약 서버", "영향 범위"), start=2):
        cell = ws.cell(start_row + 1, c, label); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = BORDER_ALL
    repeated = sorted(((e["취약"], code, e) for code, e in by_code.items() if e["취약"]), reverse=True)[:5]
    for row, (count, code, e) in enumerate(repeated, start=start_row + 2):
        vals = (code, e["title"], ", ".join(e["취약서버"]), f"{count}대")
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row, c, v); cell.border = BORDER_ALL; cell.alignment = WRAP if c in (3, 4) else CENTER
        ws.row_dimensions[row].height = _wrap_row_height([e["title"], ", ".join(e["취약서버"])], 26)

    # 중요도별 진단 결과 (반복 취약점 표 뒤에 고정 간격을 두고 이어붙임 -
    # 위 표들이 최대 5행까지 쓰므로 겹치지 않게 row 35부터 시작)
    # N/A는 "해당 없음"이라 양호율 분모/표시 모두에서 제외한다.
    r0 = 35
    ws.cell(r0, 2, "중요도별 진단 결과").font = Font(bold=True, size=11)
    imp_headers = ["중요도", "양호", "취약", "검토", "양호율"]
    for i, hd in enumerate(imp_headers):
        ws.cell(r0 + 1, 2 + i, hd)
    _style_header_row(ws, r0 + 1, len(imp_headers))
    row = r0 + 2
    for imp in ("상", "중", "하"):
        stat = by_importance.get(imp, {"양호": 0, "취약": 0, "검토": 0, "N/A": 0})
        tot = stat["양호"] + stat["취약"] + stat["검토"]
        rate = f"{(stat['양호'] / tot * 100):.1f}%" if tot else "0.0%"
        vals = [imp, stat["양호"], stat["취약"], stat["검토"], rate]
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row, c, v); cell.border = BORDER_ALL
            cell.alignment = CENTER if c > 2 else Alignment(horizontal="left", vertical="center")
        row += 1

    # 점검 영역별 결과
    r1 = row + 2
    ws.cell(r1, 2, "점검 영역별 결과").font = Font(bold=True, size=11)
    for i, hd in enumerate(imp_headers):
        ws.cell(r1 + 1, 2 + i, "점검 영역" if i == 0 else hd)
    _style_header_row(ws, r1 + 1, len(imp_headers))
    row = r1 + 2
    for cat, stat in by_category.items():
        tot = stat["양호"] + stat["취약"] + stat["검토"]
        rate = f"{(stat['양호'] / tot * 100):.1f}%" if tot else "0.0%"
        vals = [cat, stat["양호"], stat["취약"], stat["검토"], rate]
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row, c, v); cell.border = BORDER_ALL
            cell.alignment = CENTER if c > 2 else Alignment(horizontal="left", vertical="center")
        row += 1

    _autofit(ws, {"A": 3, "B": 22, "C": 30, "D": 22, "E": 12, "F": 10})
    return ws


def _build_summary(wb: Workbook, by_code: dict):
    ws = wb.create_sheet("항목별 요약")
    _nav_link(ws, "A1", "<< 대시보드", "대시보드")
    unix_n = sum(1 for c in by_code if c.startswith("U-")); db_n = sum(1 for c in by_code if c.startswith("D-"))
    ws["A2"] = f"점검항목 (UNIX {unix_n}개, DBMS {db_n}개 — 총 {len(by_code)}개)"
    ws["A2"].font = Font(bold=True, size=11)

    headers = ["No", "점검 영역", "코드", "항목명", "중요도", "전체", "취약", "취약률", "취약 서버 ID"]
    for i, hd in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=hd)
    _style_header_row(ws, 3, len(headers))

    for i, (code, e) in enumerate(sorted(by_code.items()), start=4):
        rate = f"{(e['취약'] / e['전체'] * 100):.1f}%" if e["전체"] else "0.0%"
        vals = [
            i - 3,
            e["category"],
            code,
            e["title"],
            e["importance"],
            e["전체"],
            e["취약"],
            rate,
            ", ".join(e["취약서버"]) if e["취약서버"] else "-",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = BORDER_ALL
            cell.alignment = WRAP if c in (2, 4, 9) else CENTER
        imp_style = IMPORTANCE_STYLE.get(e["importance"])
        if imp_style:
            cell = ws.cell(row=i, column=5)
            cell.fill = PatternFill("solid", fgColor=imp_style["fill"])
            cell.font = Font(color=imp_style["font"], bold=True)
        ws.row_dimensions[i].height = _wrap_row_height([e["category"], e["title"]], 48)

    _autofit(ws, {"A": 6, "B": 18, "C": 8, "D": 48, "E": 8, "F": 8, "G": 8, "H": 10, "I": 30})
    return ws


def _build_host_sheet(wb: Workbook, host: dict, comp: dict = None):
    """comp: backend/db.py::get_comparison_data()가 이 호스트(ip)에 대해 반환한
    entry 하나(또는 없으면 None) - "조치 전" 열을 채우는 데 쓴다. is_baseline이면
    (기준이 될 이전 회차가 없음) 모든 행에 "-"만 표시한다."""
    hostname = host.get("hostname", "unknown")
    ws = wb.create_sheet(hostname[:31])
    _nav_link(ws, "A1", "<< 대시보드", "대시보드")
    _nav_link(ws, "C1", "<< 항목별 요약", "항목별 요약")
    results = host.get("results", [])
    counts = {key: sum(_effective_status(r) == key for r in results) for key in ("양호", "취약", "검토", "N/A")}
    score = _host_score(host)
    info = [("대상정보", "Hostname", host.get("hostname", "-"), "OS/DB", host.get("os", "-"), "보안점수", f"{score}점"),
            (None, "IP", host.get("ip", "-"), "담당자", host.get("owner", "-"), "판정 현황", f"양호 {counts['양호']} · 취약 {counts['취약']} · 검토 {counts['검토']}")]
    for row_i, row_vals in enumerate(info, start=2):
        label0, k1, v1, k2, v2, k3, v3 = row_vals
        if label0: ws.cell(row_i, 1, label0).font = Font(bold=True, size=10); ws.cell(row_i, 1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        for col, value, is_label in ((2, k1, True), (3, v1, False), (5, k2, True), (6, v2, False), (8, k3, True), (9, v3, False)):
            cell = ws.cell(row_i, col, value); cell.font = LABEL_FONT if is_label else (Font(bold=True, size=10) if col == 9 else VALUE_FONT)
            cell.alignment = WRAP
    # 기준(이전) 회차가 있는 호스트만 "조치 전" 열을 채운다 - 최초 진단이라
    # 비교 대상이 없으면(comp가 없거나 is_baseline) 전부 "-"로 둔다.
    before_map = (comp or {}).get("before_map") if comp and not comp.get("is_baseline") else None
    # DOCX "조치 경과" 섹션과 동일하게, 기준↔현재 회차 사이 실제 경과일을
    # "조치 전" 열 바로 위에 캡션으로 보여준다 - 두 회차 날짜가 다 있을 때만.
    if before_map is not None:
        before_d, after_d = comp.get("before_scan_date"), comp.get("after_scan_date")
        elapsed = f" · {(after_d - before_d).days}일 경과" if before_d and after_d else ""
        caption = f"기준({comp['before_scan_id']}) → 현재({comp['after_scan_id']}){elapsed}"
        ws.cell(5, 5, caption).font = Font(italic=True, size=9, color="64748B")
    headers = ["점검영역", "CODE", "점검항목", "위험도", "조치 전", "판정 결과(조치 후)", "조치 상태", "현재 설정", "조치 권고 / 사유", "점수"]
    header_row = 6
    for i, hd in enumerate(headers, 1): ws.cell(header_row, i, hd)
    _style_header_row(ws, header_row, len(headers))
    for row, r in enumerate(results, start=7):
        sk = _effective_status(r)
        # before/after 비교가 있는 호스트(before_map 존재)에서는 "조치 후"(이번
        # 회차) 판정에 검토(수동확인 미확정)를 그대로 두지 않는다 - 아직 사람이
        # 확정 안 한 검토를 취약 쪽으로 넘겨서, "조치 전/후" 비교표에서 검토로
        # 남아 판단이 흐려지는 항목이 없게 한다.
        if before_map is not None and sk == "검토":
            sk = "취약"
        action = _action_status(r, sk)
        before_sk = before_map.get(r.get("code")) if before_map is not None else None
        prefix = f"[점검파일: {r['target_file']}]\n" if r.get("target_file") else ""
        evidence = r.get("evidence_description") or ""
        vals = [r.get("category", "-"), r.get("code", "-"), r.get("title", "-"), r.get("importance", "-"),
                before_sk or "-", sk, action,
                (prefix + evidence) if (prefix or evidence) else None, r.get("recommendation_text"), _score(sk)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v); cell.border = BORDER_ALL; cell.alignment = WRAP
        imp_style = IMPORTANCE_STYLE.get(r.get("importance")); st_style = STATUS_STYLE[sk]
        if imp_style: ws.cell(row, 4).fill = PatternFill("solid", fgColor=imp_style["fill"]); ws.cell(row, 4).font = Font(color=imp_style["font"], bold=True)
        bf_style = STATUS_STYLE.get(before_sk) if before_sk else None
        if bf_style: ws.cell(row, 5).fill = PatternFill("solid", fgColor=bf_style["fill"]); ws.cell(row, 5).font = Font(color=bf_style["font"], bold=True)
        ws.cell(row, 6).fill = PatternFill("solid", fgColor=st_style["fill"]); ws.cell(row, 6).font = Font(color=st_style["font"], bold=True)
        if action == "미조치": ws.cell(row, 7).font = Font(color="991B1B", bold=True)
        elif action == "재조치 필요": ws.cell(row, 7).font = Font(color="92400E", bold=True)
        elif action == "조치 완료": ws.cell(row, 7).font = Font(color="166534", bold=True)
        elif action in ("검토 필요", "수동 판정"): ws.cell(row, 7).font = Font(color="075985", bold=True)
        ws.row_dimensions[row].height = _wrap_row_height([r.get("title"), vals[7], vals[8]], 36)
    _autofit(ws, {"A": 14, "B": 9, "C": 31, "D": 9, "E": 10, "F": 15, "G": 13, "H": 36, "I": 36, "J": 8})
    ws.auto_filter.ref = f"A{header_row}:J{max(header_row + 1, ws.max_row)}"
    ws.freeze_panes = "A7"
    return ws


# ---------------------------------------------------------------------------
# 외부 공개 함수
# ---------------------------------------------------------------------------
def generate_xlsx(hosts_data, meta: dict = None, comparisons: list = None) -> bytes:
    """
    hosts_data: 기존 generate_csv()와 동일한 구조 (+선택적 os/owner/security_score_100 필드)
    meta: {"title":..., "subtitle":..., "customer":..., "period":...} 표지에 쓸 정보 (선택)
    comparisons: backend/db.py::get_comparison_data()의 반환값 그대로(선택) - 있으면
    각 호스트 시트에 기준(이전) 회차 대비 "조치 전" 열을 채운다. 없으면(None) 그
    열은 전부 "-"로 남는다(DOCX와 달리 이 인자가 없어도 시트 자체는 그대로 생성됨).
    반환: xlsx 파일의 바이트 (그대로 응답 body 나 파일로 저장)
    """
    meta = meta or {}
    comp_by_ip = {c["ip"]: c for c in (comparisons or [])}
    total, by_importance, by_category, by_code = _aggregate(hosts_data)
    unix_count = sum(1 for code in by_code if code.startswith("U-"))
    db_count = sum(1 for code in by_code if code.startswith("D-"))

    default_owner = meta.get("auditor") or "-"
    for h in hosts_data:
        if not h.get("owner"):
            h["owner"] = default_owner

    meta = {
        "title": meta.get("title", "서버 취약점 진단 상세 결과 보고서"),
        "subtitle": meta.get("subtitle", "UNIX·DBMS 통합 보안 진단"),
        "org": meta.get("org") or "HIGHFIVE SECURITY",
        "inspector": meta.get("inspector") or "",
        "customer": meta.get("customer") or "-",
        "period": meta.get("period") or datetime.now(KST).strftime("%Y-%m-%d"),
        "auditor": default_owner,
        "host_count": len(hosts_data),
        "item_count": len(by_code),
        "unix_count": unix_count,
        "db_count": db_count,
    }

    wb = Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    _build_cover(wb, meta)
    _build_dashboard(wb, hosts_data, total, by_importance, by_category, by_code)
    _build_summary(wb, by_code)
    # "항목별 요약" 다음에 이어지는 서버별 시트는 hostname 사전순으로 나열한다
    # (hosts_data 원래 순서는 DB에 등록/스캔된 순서라 시트 탭이 뒤죽박죽 보임).
    for h in sorted(hosts_data, key=lambda h: h.get("hostname") or ""):
        _build_host_sheet(wb, h, comp_by_ip.get(h.get("ip")))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# 기존 CSV 함수는 그대로 유지 (필요시 계속 사용 가능)
def generate_csv(hosts_data):
    import pandas as pd

    records = []
    for h in hosts_data:
        for r in h.get("results", []):
            records.append(
                {
                    "Hostname": h.get("hostname"),
                    "IP": h.get("ip"),
                    "Code": r.get("code"),
                    "Category": r.get("category"),
                    "Title": r.get("title"),
                    "Importance": r.get("importance"),
                    "Status": r.get("status"),
                    "Target_File": r.get("target_file"),
                    "Evidence": r.get("evidence_description"),
                    "Guide": r.get("recommendation_text"),
                }
            )
    df = pd.DataFrame(records)
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding="utf-8-sig")
    return output.getvalue()
