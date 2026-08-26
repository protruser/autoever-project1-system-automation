#!/usr/bin/env python3
"""
generate_report.py - 진단 결과 JSON들을 취합하여 대시보드용 JSON 및 Excel 보고서를 동시 생성

사용법 (기존과 동일):
  generate_report.py --raw-dir audit_reports/raw_json --out audit_reports/report.xlsx
  (동일한 폴더에 report.json 파일이 자동으로 함께 생성됩니다.)
"""

import argparse
import glob
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

# 컨트롤 노드의 시스템 타임존은 UTC라서 datetime.now()가 UTC 벽시계를 반환한다.
# 보고서/DB에 남는 진단 시각은 KST로 명시해서 기록한다.
KST = ZoneInfo("Asia/Seoul")

import pymysql

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ANTHROPIC_API_KEY 등을 프로젝트 루트 .env에서 읽어온다 (03_save_to_mysql.py와
# 동일한 방식) - 이 스크립트는 backend/config.py를 거치지 않고 단독 실행되므로
# 여기서도 직접 로드해야 한다.
_env_file = Path(__file__).resolve().parent.parent / ".env"
if _env_file.exists():
    with open(_env_file, "r", encoding="utf-8") as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

# 진단 스크립트가 만드는 raw JSON에 가끔 홑backslash(\s 등)가 남아
# 표준 JSON으로는 무효한 escape가 생긴다. 파싱 전에 보정한다.
_BAD_ESCAPE_RE = re.compile(r'\\(?!["\\/bfnrtu])')


def _load_json_lenient(text):
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return json.loads(_BAD_ESCAPE_RE.sub(r"\\\\", text))

# ==========================================
# 엑셀 스타일 상수 (기존 유지)
# ==========================================
STATUS_FILL = {
    "VULNERABLE": PatternFill("solid", fgColor="F8CBAD"),
    "GOOD": PatternFill("solid", fgColor="C6E0B4"),
    "MANUAL": PatternFill("solid", fgColor="FFE699"),
    "ERROR": PatternFill("solid", fgColor="D9D9D9"),
}
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# ==========================================
# 1. JSON 종합 보고서 생성용 헬퍼 함수
# ==========================================
def load_score_map(score_filepath="scores.json"):
    score_map = {}
    if not os.path.exists(score_filepath):
        print(f"[!] 점수 기준 파일({score_filepath})이 없습니다. 모든 배점이 0으로 처리됩니다.")
        return score_map
        
    try:
        with open(score_filepath, 'r', encoding='utf-8') as f:
            score_data = json.load(f)
            if isinstance(score_data, list):
                for item in score_data:
                    code = item.get("code")
                    score = item.get("score", 0)
                    if code:
                        score_map[code] = score
            elif isinstance(score_data, dict):
                score_map = score_data
    except Exception as e:
        print(f"[!] 점수 파일 파싱 에러: {e}")
        
    return score_map


def process_host_file(filepath, score_map):
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = _load_json_lenient(f.read())
    except Exception as e:
        print(f"[!] {filepath} 읽기 실패: {e}")
        return None

    host_info = data.get("host_info", {})
    raw_results = data.get("results", [])
    
    summary = {
        "total": len(raw_results),
        "pass": 0, "vuln": 0, "na": 0, "manual": 0,
        "max_score": 0, "deducted_score": 0
    }
    
    results = []
    for res in raw_results:
        code = res.get("code", "UNKNOWN")
        
        # 상태값 통합 매핑
        raw_status = res.get("status", "검토").upper()
        if raw_status in ["양호", "GOOD"]:
            status = "양호"
        elif raw_status in ["취약", "FAIL", "VULNERABLE"]:
            status = "취약"
        elif raw_status in ["N/A", "NA", "ERROR"]:
            status = "N/A"
        else:
            status = "검토"
            
        weight = score_map.get(code, 0)
        risk = 0
        
        if status == "양호":
            summary["pass"] += 1
            summary["max_score"] += weight
        elif status == "취약":
            summary["vuln"] += 1
            summary["max_score"] += weight
            summary["deducted_score"] += weight
            risk = weight
        elif status == "N/A":
            summary["na"] += 1
        else:
            summary["manual"] += 1
            summary["max_score"] += weight
            
        res["weight_score"] = weight
        res["risk_score"] = risk
        res["status"] = status
        results.append(res)
        
    valid_total = summary["total"] - summary["na"]
    comp_rate = (summary["pass"] / valid_total * 100) if valid_total > 0 else 100
    summary["compliance_rate"] = f"{comp_rate:.1f}%"
    
    sec_score = ((summary["max_score"] - summary["deducted_score"]) / summary["max_score"] * 100) if summary["max_score"] > 0 else 100
    summary["security_score_100"] = round(sec_score, 2)
    summary["security_score_ratio"] = round(sec_score / 100, 2)
    
    # [MOD] 미래창조과학부고시 제2013-37호 취약점 분석·평가 점수 산출식 예시의
    # 등급표(5단계, 91/81/71/61점 기준)로 일치 - backend/db.py의 등급 재계산
    # 로직과 동일한 기준을 쓴다(예전엔 여기만 3단계(80/60점) 자체 기준이었음).
    if sec_score >= 91:
        summary["grade"], summary["grade_color"] = "우수", "green"
    elif sec_score >= 81:
        summary["grade"], summary["grade_color"] = "양호", "green"
    elif sec_score >= 71:
        summary["grade"], summary["grade_color"] = "보통", "yellow"
    elif sec_score >= 61:
        summary["grade"], summary["grade_color"] = "미흡", "orange"
    else:
        summary["grade"], summary["grade_color"] = "취약", "red"
        
    summary["last_diagnosed_at"] = datetime.fromtimestamp(
        os.path.getmtime(filepath), tz=KST
    ).strftime("%Y-%m-%d %H:%M:%S")

    return {
        "host_info": host_info,
        "summary": summary,
        "results": results
    }


# ==========================================
# 2. 엑셀 보고서 생성용 헬퍼 함수
# ==========================================
def load_results(raw_dir):
    rows = []
    for path in sorted(glob.glob(os.path.join(raw_dir, "*.json"))):
        try:
            with open(path, encoding="utf-8") as f:
                data = _load_json_lenient(f.read())
        except (json.JSONDecodeError, OSError) as e:
            print(f"[skip] {path}: {e}")
            continue
            
        if isinstance(data, dict) and "results" in data:
            h_info = data.get("host_info", {})
            for item in data["results"]:
                item["hostname"] = h_info.get("hostname", "unknown")
                item["os_type"] = h_info.get("os", "")
                item["os_version"] = h_info.get("kernel", "")
                item["check_id"] = item.get("code", "")
                
                st = item.get("status", "").upper()
                if st in ["양호", "GOOD"]: item["status"] = "GOOD"
                elif st in ["취약", "FAIL", "VULNERABLE"]: item["status"] = "VULNERABLE"
                elif st in ["N/A", "NA", "ERROR"]: item["status"] = "ERROR"
                else: item["status"] = "MANUAL"
                
                item["_source_file"] = os.path.basename(path)
                rows.append(item)
        else:
            if isinstance(data, dict):
                data = [data]
            for item in data:
                item["_source_file"] = os.path.basename(path)
                rows.append(item)
    return rows


def style_header(ws, headers):
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autofit(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_summary_sheet(wb, rows):
    ws = wb.active
    ws.title = "요약"

    by_host = defaultdict(Counter)
    for r in rows:
        by_host[r.get("hostname", "unknown")][r.get("status", "ERROR")] += 1

    total = Counter(r.get("status", "ERROR") for r in rows)
    ws.append(["KISA 주요정보통신기반시설 취약점 진단 결과 요약"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])
    ws.append(["구분", "GOOD", "VULNERABLE", "MANUAL", "ERROR", "합계"])
    style_header(ws, ["구분", "GOOD", "VULNERABLE", "MANUAL", "ERROR", "합계"])

    ws.append(["전체", total["GOOD"], total["VULNERABLE"], total["MANUAL"], total["ERROR"], sum(total.values())])
    for host, c in sorted(by_host.items()):
        ws.append([host, c["GOOD"], c["VULNERABLE"], c["MANUAL"], c["ERROR"], sum(c.values())])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, int) and cell.value > 0:
                cell.fill = STATUS_FILL["VULNERABLE"]

    autofit(ws, [22, 10, 12, 10, 10, 10])


def build_detail_sheet(wb, rows):
    ws = wb.create_sheet("상세")
    headers = ["hostname", "check_id", "category", "status", "current_value",
               "expected_value", "os_type", "os_version", "timestamp"]
    ws.append(headers)
    style_header(ws, headers)

    rows_sorted = sorted(rows, key=lambda r: (r.get("hostname", ""), r.get("check_id", "")))
    for r in rows_sorted:
        ws.append([r.get(h, "") for h in headers])
        status = r.get("status", "ERROR")
        fill = STATUS_FILL.get(status)
        if fill:
            ws.cell(row=ws.max_row, column=headers.index("status") + 1).fill = fill

    autofit(ws, [18, 10, 20, 12, 40, 30, 10, 12, 20])


def build_manual_sheet(wb, rows):
    ws = wb.create_sheet("수동조치 필요")
    headers = ["hostname", "check_id", "category", "current_value", "expected_value"]
    ws.append(headers)
    style_header(ws, headers)
    for r in sorted(rows, key=lambda r: (r.get("hostname", ""), r.get("check_id", ""))):
        if r.get("status") != "MANUAL":
            continue
        ws.append([r.get(h, "") for h in headers])
    autofit(ws, [18, 10, 20, 40, 30])


# ==========================================
# 2.5. AI 종합 소견(consultant_comment) 생성
# ==========================================
def generate_consultant_comment(final_report):
    """스캔 요약을 Claude에게 넘겨 한국어 종합 소견을 짧게 생성한다.

    audit_scans.consultant_comment는 컬럼도 있고 03_save_to_mysql.py가 저장까지
    하는데, 정작 이 값을 채우는 코드가 없어서 지금까지 모든 회차가 항상 빈
    문자열이었다 - 그 자리를 채우는 함수.

    API 키가 없거나(.env에 ANTHROPIC_API_KEY 미설정), anthropic 패키지가 없거나,
    호출이 실패해도 예외를 밖으로 던지지 않는다 - 이 스크립트는 리포트 생성/DB
    적재 파이프라인의 필수 단계라, AI 코멘트 하나 때문에 전체 스캔 저장이
    실패하면 안 된다. 실패 시 빈 문자열을 반환한다.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        print("[i] ANTHROPIC_API_KEY가 없어 종합 소견 생성을 건너뜁니다.")
        return ""

    try:
        import anthropic
    except ImportError:
        print("[i] anthropic 패키지가 없어 종합 소견 생성을 건너뜁니다 (pip install anthropic).")
        return ""

    t = final_report["total_summary"]
    lines = [
        f"진단 대상: {t['total_hosts']}대 서버, 총 {t['total_checks']}개 항목",
        f"양호 {t['total_pass']} / 취약 {t['total_vuln']} / N-A {t['total_na']}",
        f"평균 보안 점수: {t['average_security_score']}점 ({t['total_grade']})",
        "",
        "서버별 요약:",
    ]
    for h in final_report["hosts"]:
        hi, hs = h["host_info"], h["summary"]
        lines.append(f"- {hi.get('hostname')} ({hi.get('os')}): {hs.get('security_score_100')}점, 취약 {hs.get('vuln')}건")

    # 프롬프트가 너무 커지지 않도록, 우선순위 파악에 실제로 쓰이는 '중요도 상'
    # 취약 항목만 추린다 (전체 결과를 다 넣을 필요는 없음).
    critical = []
    for h in final_report["hosts"]:
        hostname = h["host_info"].get("hostname")
        for r in h["results"]:
            if r.get("status") == "취약" and r.get("importance") == "상":
                critical.append(f"- [{hostname}] {r.get('code')} {r.get('title')}")
    if critical:
        lines.append("")
        lines.append("중요도 '상' 취약 항목:")
        lines.extend(critical[:20])

    summary_text = "\n".join(lines)

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-opus-5",
            max_tokens=1024,
            system=(
                "당신은 KISA 주요정보통신기반시설 취약점 진단 보고서를 작성하는 보안 컨설턴트입니다. "
                "아래 진단 요약만 보고, 이번 회차의 전반적인 보안 수준과 우선 조치가 필요한 부분을 "
                "한국어 3~5문장으로 간결하게 작성하세요. 요약에 없는 내용은 추측해서 쓰지 말고, "
                "수치를 인용할 땐 요약에 있는 값만 그대로 쓰세요."
            ),
            messages=[{"role": "user", "content": summary_text}],
        )
        text = "".join(b.text for b in response.content if b.type == "text").strip()
        if text:
            print("[+] 종합 소견 생성 완료")
        return text
    except Exception as e:
        print(f"[!] 종합 소견 생성 실패(무시하고 진행): {e}")
        return ""


# ==========================================
# 2.6. scan_id 회차 번호 조회
# ==========================================
def next_run_number(db_name, date_str):
    """오늘(date_str, YYYYMMDD) 날짜로 이미 등록된 scan_id들 중 가장 큰 회차
    번호 다음 값을 반환한다(없으면 1). db_name이 없거나 DB 연결에 실패하면
    (최초 실행, DB 아직 없음 등) 예외를 밖으로 던지지 않고 1로 폴백한다 -
    이 스크립트는 리포트 생성 파이프라인의 필수 단계라 회차 번호 하나
    때문에 전체가 실패하면 안 된다."""
    if not db_name:
        return 1
    try:
        conn = pymysql.connect(
            host=os.getenv("AUDIT_DB_HOST", "localhost"),
            port=int(os.getenv("AUDIT_DB_PORT", 3306)),
            user=os.getenv("DB_APP_USER", "audit_user"),
            password=os.getenv("DB_APP_PASSWORD", ""),
            database=db_name,
            charset="utf8mb4",
        )
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT scan_id FROM audit_scans WHERE scan_id LIKE %s",
                    (f"SCAN-{date_str}-%",)
                )
                rows = cur.fetchall()
        finally:
            conn.close()

        max_n = 0
        for (scan_id,) in rows:
            try:
                max_n = max(max_n, int(scan_id.rsplit("-", 1)[-1]))
            except ValueError:
                continue
        return max_n + 1
    except Exception as e:
        print(f"[i] scan_id 회차 번호 조회 실패({e}) - 1회차로 진행합니다.")
        return 1


# ==========================================
# 3. 메인 실행부
# ==========================================
def main():
    ap = argparse.ArgumentParser()
    # 기존과 동일한 옵션 체계 유지
    ap.add_argument("--raw-dir", default="audit_reports/raw_json", help="진단 결과 JSON 파일들이 있는 디렉토리")
    ap.add_argument("--out", default="audit_reports/report.xlsx", help="저장될 통합 엑셀 보고서 경로")
    ap.add_argument("--score-file", default="scores.json", help="항목별 배점 기준 JSON 파일")
    # [MOD] scan_id를 하루에 한 번(-01 고정)이 아니라 실제 회차 수만큼 늘리기
    # 위해 대상 DB를 받는다 - 오늘 이미 등록된 scan_id 개수를 조회하는 데 씀
    # (같은 날 여러 번 스캔해도 회차별로 before/after 비교가 되게 하려는 목적).
    ap.add_argument("--db", default=None, help="회차 번호 조회에 쓸 대상 DB명(예: audit_autoever_2026). 생략 시 항상 1회차로 진행")
    args = ap.parse_args()

    # --out 인자(report.xlsx)를 바탕으로 JSON 파일(report.json) 이름과 경로 자동 생성
    out_dir = os.path.dirname(args.out) or "."
    out_name = os.path.splitext(os.path.basename(args.out))[0]
    out_json = os.path.join(out_dir, f"{out_name}.json")

    # [STEP 1] JSON 종합 보고서 생성 로직
    score_map = load_score_map(args.score_file)
    host_files = sorted(glob.glob(os.path.join(args.raw_dir, "*.json")))
    
    if not host_files:
        print(f"[!] '{args.raw_dir}' 에서 진단 결과 JSON을 찾지 못했습니다.")
        return

    hosts_data = []
    total = {"hosts": 0, "checks": 0, "pass": 0, "vuln": 0, "na": 0, "max_score": 0, "deducted": 0}
    
    for path in host_files:
        host_data = process_host_file(path, score_map)
        if host_data:
            hosts_data.append(host_data)
            total["hosts"] += 1
            total["checks"] += host_data["summary"]["total"]
            total["pass"] += host_data["summary"]["pass"]
            total["vuln"] += host_data["summary"]["vuln"]
            total["na"] += host_data["summary"]["na"]
            total["max_score"] += host_data["summary"]["max_score"]
            total["deducted"] += host_data["summary"]["deducted_score"]

    valid_checks = total["checks"] - total["na"]
    avg_comp = (total["pass"] / valid_checks * 100) if valid_checks > 0 else 100
    avg_sec = ((total["max_score"] - total["deducted"]) / total["max_score"] * 100) if total["max_score"] > 0 else 100
    
    # [MOD] 위와 동일한 이유 - 공식 5단계 등급표로 일치
    total_grade, total_color = "우수", "green"
    if avg_sec < 91: total_grade, total_color = "양호", "green"
    if avg_sec < 81: total_grade, total_color = "보통", "yellow"
    if avg_sec < 71: total_grade, total_color = "미흡", "orange"
    if avg_sec < 61: total_grade, total_color = "취약", "red"

    # [MOD] 하루에 한 번(-01 고정)이 아니라, 오늘 이미 몇 회차인지 DB에서 확인해
    # 다음 번호를 매긴다 - 같은 날 여러 번 스캔해도 회차가 안 뭉개지고 각각
    # 별도 scan_id(before/after 비교 단위)로 남는다.
    _now_kst = datetime.now(KST)
    _date_str = _now_kst.strftime("%Y%m%d")
    _run_n = next_run_number(args.db, _date_str)

    final_report = {
        "scan_info": {
            "scan_id": f"SCAN-{_date_str}-{_run_n:02d}",
            "project_name": "주요정보통신기반시설 시스템 취약점 진단",
            "scan_date": _now_kst.strftime("%Y-%m-%d %H:%M:%S"),
            "auditor": "심수용, 김성진, 김하영, 정진우, 한주협"
        },
        "total_summary": {
            "total_hosts": total["hosts"],
            "total_checks": total["checks"],
            "total_pass": total["pass"],
            "total_vuln": total["vuln"],
            "total_na": total["na"],
            "average_compliance_rate": f"{avg_comp:.1f}%",
            "average_security_score": round(avg_sec, 2),
            "average_security_ratio": round(avg_sec / 100, 2),
            "total_grade": total_grade,
            "total_grade_color": total_color
        },
        "hosts": hosts_data
    }

    final_report["scan_info"]["consultant_comment"] = generate_consultant_comment(final_report)

    os.makedirs(out_dir, exist_ok=True)
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(final_report, f, ensure_ascii=False, indent=2)
    print(f"[+] 통합 JSON 생성 완료: {out_json} (호스트 {len(hosts_data)}대 취합)")

    # [STEP 2] Excel 보고서 생성 로직[cite: 4]
    rows = load_results(args.raw_dir)
    if rows:
        wb = Workbook()
        build_summary_sheet(wb, rows)
        build_detail_sheet(wb, rows)
        build_manual_sheet(wb, rows)

        wb.save(args.out)
        print(f"[+] 엑셀 보고서 생성 완료: {args.out} ({len(rows)}건)")

    # [STEP 3] raw_json 정리 - 이번 회차에 실제로 읽은 파일들을 raw_json_old/로
    # 옮긴다. 안 하면 다음 실행 때도 이 파일들이 다시 글롭에 걸려서, 이번
    # 회차뿐 아니라 앞으로의 모든 회차 종합점수/등급이 과거 스냅샷 누적으로
    # 계속 오염된다(실측됨: raw_json/에 8/24부터 59개 파일이 안 지워지고
    # 쌓여서, 실제 호스트 5대가 아니라 58개 항목으로 평균을 내고 있었음).
    # 삭제가 아니라 이동인 이유: 원본 진단 근거는 보존하면서 다음 글롭
    # 대상에서만 빼기 위함.
    old_dir = os.path.join(args.raw_dir, "..", "raw_json_old")
    old_dir = os.path.normpath(old_dir)
    os.makedirs(old_dir, exist_ok=True)
    moved = 0
    for path in host_files:
        try:
            dest = os.path.join(old_dir, os.path.basename(path))
            if os.path.exists(dest):
                # 같은 파일명이 이미 old에 있으면(재실행 등) 타임스탬프를 붙여 보존
                base, ext = os.path.splitext(dest)
                dest = f"{base}.{int(datetime.now(KST).timestamp())}{ext}"
            os.replace(path, dest)
            moved += 1
        except OSError as e:
            print(f"[!] {path} 정리 실패(다음 실행 때 다시 읽힘): {e}")
    print(f"[+] raw_json 정리 완료: {moved}개 파일을 {old_dir}로 이동")


if __name__ == "__main__":
    main()
